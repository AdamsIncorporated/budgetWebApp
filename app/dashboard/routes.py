from sqlalchemy import func
from flask import (
    Blueprint,
    render_template,
    redirect,
    flash,
    url_for,
    request,
    abort,
    session,
)
from flask_login import login_required, current_user
import base64
from io import BytesIO
from app import db
from repositories.models import (
    MasterEmail,
    BusinessUnit,
    UserBusinessUnit,
    User,
    BudgetEntryAdminView,
    Account,
)
from .forms import (
    MasterEmailForm,
    UserBusinessUnits,
    MultiviewTemplate,
    BudgetEntryAdminViewCreateForm,
    BudgetEntryAdminViewsForm,
    DashBoardActualsToBudgetForm,
)
from flask import jsonify, current_app, send_file
from sqlalchemy import text
from repositories.queries import (
    queries,
    get_default_historical_fiscal_year,
    get_default_business_unit,
)
from services.current_user_image import image_wrapper
from services.get_current_fiscal_year import get_fiscal_year
from openpyxl import load_workbook
import os


dashboard = Blueprint(
    "dashboard",
    __name__,
    template_folder="templates/dashboard",
    static_folder="static",
    url_prefix="/dashboard",
)


@dashboard.route("/home/table")
def get_table_html():
    fiscal_year = request.args.get("fiscalYear")
    business_unit_id = request.args.get("businessUnitId")
    query = queries["actual_by_budget_for_fiscal_year_and_business_unit"](
        fiscal_year, business_unit_id
    )
    data = db.session.execute(text(query)).fetchall()

    return render_template("homeComponents/table.html", data=data)


@dashboard.route("/home")
@login_required
@image_wrapper
def home(image_file=None):
    form = DashBoardActualsToBudgetForm()

    return render_template("homeDashboard.html", image_file=image_file, form=form)


@dashboard.route("/download-template/<string:fiscal_year>")
@login_required
def download_template(fiscal_year):
    download_folder = os.path.join(current_app.root_path, r"dashboard\download")

    # Ensure the download folder exists
    if not os.path.exists(download_folder):
        raise FileNotFoundError(f"The directory '{download_folder}' does not exist.")

    # Define the Excel file path
    excel_file_path = os.path.join(download_folder, "template.xlsx")
    workbook = load_workbook(excel_file_path)

    # Get the active sheet
    sheet = workbook.active
    sheet["B2"] = fiscal_year

    # get the data from the server
    query = queries["multiview_download"]
    data = db.session.execute(text(query), {"proposed_fy": fiscal_year}).all()

    for row_idx, row in enumerate(data, start=9):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Optional: Return the file for download as a response
    return send_file(
        output,
        as_attachment=True,
        download_name="template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@dashboard.route("/template", methods=["GET", "POST"])
@login_required
@image_wrapper
def template(image_file=None):
    form = MultiviewTemplate()
    current_fiscal_year = get_fiscal_year()

    if request.method == "GET":
        query = queries["multiview_download"]
        form.fiscal_year.data = current_fiscal_year
        results = db.session.execute(
            text(query), {"proposed_fy": current_fiscal_year}
        ).all()

    if request.method == "POST":
        if form.validate_on_submit():
            new_fiscal_year = form.fiscal_year.data
            query = queries["multiview_download"]
            results = db.session.execute(
                text(query), {"proposed_fy": new_fiscal_year}
            ).all()
            form.fiscal_year.data = new_fiscal_year

    return render_template(
        "template.html", image_file=image_file, form=form, results=results
    )


@dashboard.route("/account-actuals-timeline")
def account_actuals_timeline():
    result = (
        db.session.execute(text(queries["account_actuals_timeline"])).mappings().all()
    )
    result_dicts = [dict(row) for row in result]
    return jsonify(records=result_dicts)


@dashboard.route("/budget-pie-chart")
def budget_pie_chart():
    result = db.session.execute(text(queries["budget_pie_chart"])).mappings().all()
    result_dicts = [dict(row) for row in result]
    return jsonify(records=result_dicts)


@dashboard.route("/add-users")
@login_required
def add_users():
    master_emails = MasterEmail.query.all()
    image_file = None
    if current_user.image_file:
        image_file = base64.b64encode(current_user.image_file).decode("utf-8")

    return render_template(
        "addUsers.html", master_emails=master_emails, image_file=image_file
    )


@dashboard.route("/create", methods=["GET", "POST"])
@login_required
def create():
    form = MasterEmailForm()

    if request.method == "GET":
        query = (
            db.session.query(BusinessUnit.business_unit_id, BusinessUnit.business_unit)
            .distinct()
            .order_by(BusinessUnit.business_unit_id)
        )
        results = query.all()
        data = {"user_business_units": results}
        form.process(data=data)

    if request.method == "POST":
        if form.validate_on_submit():
            new_email = MasterEmail(
                email=form.email.data,
                user_creator_id=current_user.id,
            )
            db.session.add(new_email)
            db.session.commit()

            # query users to get id
            user_id = User.query.filter_by(email=form.email.data).first_or_404().id

            for row in form.user_business_units.data:
                if row.get("is_business_unit_selected"):
                    business_unit_id = int(row.get("business_unit_id"))
                    new_user_business_unit = UserBusinessUnit(
                        user_id=user_id, business_unit_id=business_unit_id
                    )
                    db.session.add(new_user_business_unit)
            db.session.commit()

            flash(f"User rights for {form.email.data} created successfully!", "success")
            return redirect(url_for("dashboard.add_users"))
        else:
            errors = ". ".join(
                [
                    error
                    for field_errors in form.errors.values()
                    for error in field_errors
                ]
            )
            flash(errors, "error")
            return redirect(url_for("dashboard.add_users"))

    return render_template("modal/createModal.html", form=form)


@dashboard.route("/edit", methods=["GET", "POST"])
@login_required
def edit():
    form = UserBusinessUnits()

    if request.method == "GET":
        user_id = int(request.args.get("id"))
        master_email = MasterEmail.query.filter_by(id=user_id).first()

        query = queries["user_business_units"](user_id=user_id)
        user_business_units = db.session.execute(text(query)).fetchall()

        data = {
            "user_id": user_id,
            "email": master_email.email,
            "date_created": master_email.date_created,
            "user_business_units": [
                {
                    "id": id,
                    "business_unit_id": business_unit_id,
                    "business_unit": business_unit,
                    "is_business_unit_selected": is_business_unit_selected,
                }
                for id, business_unit_id, business_unit, is_business_unit_selected in user_business_units
            ],
        }
        form.process(data=data)

    if request.method == "POST":
        if form.validate_on_submit():
            user_id = int(form.user_id.data)
            current_units = UserBusinessUnit.query.filter_by(user_id=user_id).all()
            current_unit_ids = {(ub.id, ub.business_unit_id) for ub in current_units}

            selected_units = {
                (row["id"], row["business_unit_id"])
                for row in form.user_business_units.data
                if row["is_business_unit_selected"]
            }
            units_to_delete = current_unit_ids - selected_units
            units_to_add = selected_units - current_unit_ids

            # Delete unselected business units
            if units_to_delete:
                for unit_id, business_unit_id in units_to_delete:
                    UserBusinessUnit.query.filter_by(
                        id=unit_id, user_id=user_id
                    ).delete(synchronize_session=False)

            # Add new selected business units
            for unit_id, business_unit_id in units_to_add:
                new_unit = UserBusinessUnit(
                    user_id=user_id, business_unit_id=business_unit_id
                )
                db.session.add(new_unit)

            db.session.commit()

            flash(f"Successfully updated {form.email.data} user!", "success")

            return redirect(url_for("dashboard.add_users"))
        else:
            errors = ". ".join(
                [
                    error
                    for field_errors in form.errors.values()
                    for error in field_errors
                ]
            )
            flash(errors, "error")
            return redirect(url_for("dashboard.add_users"))

    return render_template("modal/editModal.html", form=form)


@dashboard.route("/delete", methods=["GET", "POST"])
@login_required
def delete():
    id = int(request.args.get("id"))
    data = MasterEmail.query.filter_by(id=id).first()

    if request.method == "POST":
        master_email = MasterEmail.query.get(id)

        if data:
            db.session.delete(master_email)
            db.session.commit()
            flash("Successfully deleted user!", "success")

            return redirect(url_for("dashboard.add_users"))

        else:
            flash("User not found!", "error")

            return redirect(url_for("dashboard.add_users"))

    return render_template("modal/deleteModal.html", data=data)


@dashboard.route("/budget-admin-view", methods=["GET", "POST"])
@login_required
@image_wrapper
def budget_admin_view(image_file=None):
    form = BudgetEntryAdminViewsForm()
    modal_html = request.args.get("modal_html", default=None)

    if request.method == "GET":
        data = BudgetEntryAdminView.query.all()
        form.process(data={"budget_entries": data})

    if request.method == "POST":
        if form.validate_on_submit():
            for sub_form in form.budget_entries:
                if sub_form.data["is_updated"] == "yes":
                    id = int(sub_form.data["id"])
                    row = BudgetEntryAdminView.query.filter_by(id=id).first_or_404()
                    row.display_order = sub_form.display_order.data
                    row.forcast_multiplier = sub_form.forecast_multiplier.data
                    row.forecast_comments = sub_form.forecast_comments.data
                    db.session.commit()
            flash("Successfully updated!", "success")
        else:
            flash("Form errors", "error")

    return render_template(
        "budgetAdminView.html", form=form, image_file=image_file, modal_html=modal_html
    )


@dashboard.route("/budget-admin-view/delete", methods=["GET", "POST"])
@login_required
def budget_admin_view_delete():
    id = int(request.args.get("id"))
    data = BudgetEntryAdminView.query.filter_by(id=id).first()

    if request.method == "POST":
        row = BudgetEntryAdminView.query.get(id)

        if data:
            db.session.delete(row)
            db.session.commit()
            flash("Successfully deleted row!", "success")

            return redirect(url_for("dashboard.budget_admin_view"))

        else:
            flash("Error in deleting row!", "error")

            return redirect(url_for("dashboard.budget_admin_view"))

    return render_template("budgetAdminViewModal/deleteModal.html", data=data)


@dashboard.route("/budget-admin-view/create", methods=["GET", "POST"])
@login_required
def budget_admin_view_create():
    form = BudgetEntryAdminViewCreateForm()

    if request.method == "POST":
        if form.validate_on_submit():
            account_no = (
                Account.query.filter_by(account=form.account.data)
                .with_entities(Account.account_no)
                .first()
            )[0]
            display_order = (
                BudgetEntryAdminView.query.with_entities(
                    func.max(BudgetEntryAdminView.display_order)
                ).scalar()
                + 1
            )
            new_row = BudgetEntryAdminView(
                display_order=display_order,
                account_no=account_no,
                account=form.account.data,
                rad=form.rad.data,
                forecast_multiplier=form.forecast_multiplier.data,
                forecast_comments=form.forecast_comments.data,
            )
            db.session.add(new_row)

            try:
                db.session.commit()
            except Exception as error:
                raise error

            flash(f"New budget entry admin view row created successfully!", "success")
            return redirect(url_for("dashboard.budget_admin_view"))

    return render_template("budgetAdminViewModal/createModal.html", form=form)


@dashboard.route("/get-rads")
def get_rads():
    account = request.args.get("Account", default=None, type=str)

    if not account:
        abort(404)

    query = queries["fetch_rads_by_account"](account)
    result = db.session.execute(text(query))
    options = [(item[0]) for item in result]

    return jsonify(options)
